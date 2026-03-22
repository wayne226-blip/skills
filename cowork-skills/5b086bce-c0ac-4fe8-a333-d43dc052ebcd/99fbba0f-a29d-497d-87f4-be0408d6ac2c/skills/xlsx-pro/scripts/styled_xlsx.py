#!/usr/bin/env python3
"""Create professionally styled Excel files from DataFrames using Pandas Styler.

Provides a fast path for creating presentation-quality spreadsheets with
conditional formatting, color scales, and polished headers — much less code
than raw openpyxl for report-style output.

Usage as CLI:
    python styled_xlsx.py input.csv output.xlsx
    python styled_xlsx.py input.csv output.xlsx --sheet "Sales Data"
    python styled_xlsx.py input.csv output.xlsx --theme blue --autofilter

Usage as library:
    from styled_xlsx import StyledExcelWriter

    writer = StyledExcelWriter("output.xlsx", theme="blue")
    writer.add_sheet(df, "Revenue", money_cols=["Revenue", "Profit"])
    writer.add_sheet(df2, "Metrics", pct_cols=["Growth", "Margin"], highlight_max=True)
    writer.save()
"""

import sys
import argparse
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Pre-defined color themes
THEMES = {
    "blue": {
        "header_bg": "1A365D",
        "header_fg": "FFFFFF",
        "accent": "2B6CB0",
        "alt_row": "EBF8FF",
        "border": "E2E8F0",
        "positive": "38A169",
        "negative": "E53E3E",
    },
    "green": {
        "header_bg": "1C4532",
        "header_fg": "FFFFFF",
        "accent": "38A169",
        "alt_row": "F0FFF4",
        "border": "E2E8F0",
        "positive": "38A169",
        "negative": "E53E3E",
    },
    "dark": {
        "header_bg": "1A202C",
        "header_fg": "FFFFFF",
        "accent": "4A5568",
        "alt_row": "F7FAFC",
        "border": "CBD5E0",
        "positive": "48BB78",
        "negative": "FC8181",
    },
    "minimal": {
        "header_bg": "F7FAFC",
        "header_fg": "1A202C",
        "accent": "718096",
        "alt_row": "FFFFFF",
        "border": "E2E8F0",
        "positive": "38A169",
        "negative": "E53E3E",
    },
    "corporate": {
        "header_bg": "002060",
        "header_fg": "FFFFFF",
        "accent": "0070C0",
        "alt_row": "DAEEF3",
        "border": "BDD7EE",
        "positive": "00B050",
        "negative": "FF0000",
    },
}


class StyledExcelWriter:
    """Create professional Excel files with styled DataFrames."""

    def __init__(self, output_path, theme="blue"):
        self.output_path = output_path
        self.theme = THEMES.get(theme, THEMES["blue"])
        self.sheets = []

    def add_sheet(
        self,
        df,
        sheet_name="Sheet1",
        money_cols=None,
        pct_cols=None,
        int_cols=None,
        highlight_max=False,
        highlight_min=False,
        color_negative=True,
        autofilter=True,
        freeze_panes=True,
        col_widths=None,
    ):
        """Add a styled DataFrame sheet.

        Args:
            df: pandas DataFrame
            sheet_name: Name of the sheet
            money_cols: Columns to format as currency ($#,##0)
            pct_cols: Columns to format as percentage (0.0%)
            int_cols: Columns to format as integers (#,##0)
            highlight_max: Highlight max value in numeric columns
            highlight_min: Highlight min value in numeric columns
            color_negative: Color negative numbers red
            autofilter: Add autofilter to headers
            freeze_panes: Freeze the header row
            col_widths: Dict of {column_name: width} overrides
        """
        self.sheets.append(
            {
                "df": df,
                "name": sheet_name,
                "money_cols": money_cols or [],
                "pct_cols": pct_cols or [],
                "int_cols": int_cols or [],
                "highlight_max": highlight_max,
                "highlight_min": highlight_min,
                "color_negative": color_negative,
                "autofilter": autofilter,
                "freeze_panes": freeze_panes,
                "col_widths": col_widths or {},
            }
        )

    def save(self):
        """Write all sheets to the Excel file with styling."""
        if not self.sheets:
            raise ValueError("No sheets added. Use add_sheet() first.")

        # Write raw data first
        with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
            for sheet_info in self.sheets:
                sheet_info["df"].to_excel(
                    writer, sheet_name=sheet_info["name"], index=False
                )

        # Now style it
        wb = load_workbook(self.output_path)
        t = self.theme

        header_font = Font(name="Arial", bold=True, color=t["header_fg"], size=11)
        header_fill = PatternFill("solid", fgColor=t["header_bg"])
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        body_font = Font(name="Arial", size=10)
        alt_fill = PatternFill("solid", fgColor=t["alt_row"])
        no_fill = PatternFill("solid", fgColor="FFFFFF")

        thin_border = Border(
            bottom=Side(style="thin", color=t["border"]),
        )

        pos_font = Font(name="Arial", size=10, color=t["positive"])
        neg_font = Font(name="Arial", size=10, color=t["negative"])

        for sheet_info in self.sheets:
            ws = wb[sheet_info["name"]]
            df = sheet_info["df"]
            num_rows = len(df)
            num_cols = len(df.columns)

            # Style headers
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = Border(
                    bottom=Side(style="medium", color=t["accent"])
                )

            # Style body rows
            for row_idx in range(2, num_rows + 2):
                is_alt = (row_idx - 2) % 2 == 1
                for col_idx in range(1, num_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = body_font
                    cell.fill = alt_fill if is_alt else no_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

            # Number formatting
            col_names = list(df.columns)
            for col_name in sheet_info["money_cols"]:
                if col_name in col_names:
                    col_idx = col_names.index(col_name) + 1
                    for row_idx in range(2, num_rows + 2):
                        ws.cell(row=row_idx, column=col_idx).number_format = (
                            '$#,##0;($#,##0);"-"'
                        )

            for col_name in sheet_info["pct_cols"]:
                if col_name in col_names:
                    col_idx = col_names.index(col_name) + 1
                    for row_idx in range(2, num_rows + 2):
                        ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"

            for col_name in sheet_info["int_cols"]:
                if col_name in col_names:
                    col_idx = col_names.index(col_name) + 1
                    for row_idx in range(2, num_rows + 2):
                        ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"

            # Color negative values
            if sheet_info["color_negative"]:
                for col_idx in range(1, num_cols + 1):
                    if df.dtypes.iloc[col_idx - 1] in ["float64", "int64", "float32", "int32"]:
                        for row_idx in range(2, num_rows + 2):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < 0:
                                cell.font = neg_font

            # Highlight max/min
            if sheet_info["highlight_max"] or sheet_info["highlight_min"]:
                max_fill = PatternFill("solid", fgColor="C6F6D5")
                min_fill = PatternFill("solid", fgColor="FED7D7")
                for col_idx in range(1, num_cols + 1):
                    if df.dtypes.iloc[col_idx - 1] in ["float64", "int64", "float32", "int32"]:
                        values = []
                        for row_idx in range(2, num_rows + 2):
                            val = ws.cell(row=row_idx, column=col_idx).value
                            if val is not None and isinstance(val, (int, float)):
                                values.append((row_idx, val))
                        if values:
                            if sheet_info["highlight_max"]:
                                max_row = max(values, key=lambda x: x[1])[0]
                                ws.cell(row=max_row, column=col_idx).fill = max_fill
                            if sheet_info["highlight_min"]:
                                min_row = min(values, key=lambda x: x[1])[0]
                                ws.cell(row=min_row, column=col_idx).fill = min_fill

            # Auto column widths
            for col_idx in range(1, num_cols + 1):
                col_name = col_names[col_idx - 1]
                if col_name in sheet_info["col_widths"]:
                    width = sheet_info["col_widths"][col_name]
                else:
                    # Auto-calculate: max of header and data widths
                    max_len = len(str(col_name))
                    for row_idx in range(2, min(num_rows + 2, 52)):  # Sample first 50 rows
                        val = ws.cell(row=row_idx, column=col_idx).value
                        if val is not None:
                            max_len = max(max_len, len(str(val)))
                    width = min(max_len + 3, 40)  # Cap at 40
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Row height for header
            ws.row_dimensions[1].height = 28

            # Autofilter
            if sheet_info["autofilter"]:
                ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"

            # Freeze panes
            if sheet_info["freeze_panes"]:
                ws.freeze_panes = "A2"

        wb.save(self.output_path)
        file_size = os.path.getsize(self.output_path)
        print(f"Excel created: {self.output_path} ({file_size:,} bytes)")


def main():
    parser = argparse.ArgumentParser(
        description="Create styled Excel file from CSV/TSV data"
    )
    parser.add_argument("input", help="Input CSV or TSV file")
    parser.add_argument("output", help="Output XLSX file")
    parser.add_argument("--sheet", default="Sheet1", help="Sheet name")
    parser.add_argument(
        "--theme",
        default="blue",
        choices=list(THEMES.keys()),
        help="Color theme (default: blue)",
    )
    parser.add_argument("--autofilter", action="store_true", help="Add autofilter")
    parser.add_argument(
        "--money", nargs="*", default=[], help="Columns to format as currency"
    )
    parser.add_argument(
        "--pct", nargs="*", default=[], help="Columns to format as percentage"
    )
    parser.add_argument(
        "--highlight-max", action="store_true", help="Highlight max values"
    )
    args = parser.parse_args()

    # Read input
    sep = "\t" if args.input.endswith((".tsv", ".tab")) else ","
    df = pd.read_csv(args.input, sep=sep)

    writer = StyledExcelWriter(args.output, theme=args.theme)
    writer.add_sheet(
        df,
        sheet_name=args.sheet,
        money_cols=args.money,
        pct_cols=args.pct,
        highlight_max=args.highlight_max,
        autofilter=args.autofilter,
    )
    writer.save()


if __name__ == "__main__":
    main()
